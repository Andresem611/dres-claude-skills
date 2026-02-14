#!/usr/bin/env python3
"""
Thoven Burn Rate Tracker Operations

Core functions for managing the Thoven burn rate spreadsheet.
"""

from openpyxl import load_workbook
from datetime import datetime, timedelta
import os

TRACKER_PATH = "/Users/andresmartinez/Vaults/Executive Assistant/Thoven_Burn_Rate_Tracker.xlsx"

# Known spreadsheet locations
CASH_BALANCE_CELL = "B3"
BURN_RATE_CELL = "B4"
RUNWAY_CELL = "B5"

FUNDING_START_ROW = 11
BUDGET_START_ROW = 17
RECENT_EXPENSES_START_ROW = 29

BUDGET_CATEGORIES = {
    "Meta Ads": 17,
    "Google Ads": 18,
    "Claude Code": 19,
    "Metabase": 20,
    "Canva": 21,
    "Contractors": 22,
    "Travel": 23,
    "Other SaaS": 24,
    "Teacher Onboarding": 25,
}


def load_tracker():
    """Load the burn rate tracker workbook."""
    if not os.path.exists(TRACKER_PATH):
        raise FileNotFoundError(f"Tracker not found at {TRACKER_PATH}")
    return load_workbook(TRACKER_PATH)


def get_monday_of_week():
    """Return Monday of current week in YYYY-MM-DD format."""
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    return monday.strftime("%Y-%m-%d")


def add_expense(category, amount, description=""):
    """Add expense to Recent Expenses section."""
    wb = load_tracker()
    ws = wb.active

    # Validate category
    if category not in BUDGET_CATEGORIES:
        wb.close()
        raise ValueError(f"Unknown category: {category}. Valid: {', '.join(BUDGET_CATEGORIES.keys())}")

    # Find next empty row
    row = RECENT_EXPENSES_START_ROW
    while ws[f'A{row}'].value:
        row += 1

    # Add expense
    ws[f'A{row}'] = datetime.now().strftime("%Y-%m-%d")
    ws[f'B{row}'] = category
    ws[f'C{row}'] = description or f"{category} expense"
    ws[f'D{row}'] = amount
    ws[f'D{row}'].number_format = '$#,##0'
    ws[f'E{row}'] = get_monday_of_week()

    # Update cash balance
    old_balance = ws[CASH_BALANCE_CELL].value or 0
    ws[CASH_BALANCE_CELL] = old_balance - amount

    # Get current category spend
    cat_row = BUDGET_CATEGORIES[category]
    current_spend = ws[f'C{cat_row}'].value or 0
    ws[f'C{cat_row}'] = current_spend + amount

    wb.save(TRACKER_PATH)
    wb.close()

    return {
        "category": category,
        "amount": amount,
        "description": description,
        "new_balance": old_balance - amount,
        "category_total": current_spend + amount
    }


def add_funding(source, amount, notes=""):
    """Add funding to Funding Log section."""
    wb = load_tracker()
    ws = wb.active

    # Find next empty row
    row = FUNDING_START_ROW
    while ws[f'A{row}'].value:
        row += 1

    # Add funding
    ws[f'A{row}'] = datetime.now().strftime("%Y-%m-%d")
    ws[f'B{row}'] = source
    ws[f'C{row}'] = amount
    ws[f'C{row}'].number_format = '$#,##0'
    ws[f'D{row}'] = notes

    # Update cash balance
    old_balance = ws[CASH_BALANCE_CELL].value or 0
    ws[CASH_BALANCE_CELL] = old_balance + amount

    wb.save(TRACKER_PATH)
    wb.close()

    return {
        "source": source,
        "amount": amount,
        "old_balance": old_balance,
        "new_balance": old_balance + amount
    }


def get_status():
    """Get current financial status."""
    wb = load_tracker()
    ws = wb.active

    cash = ws[CASH_BALANCE_CELL].value or 0
    burn = ws[BURN_RATE_CELL].value or 0
    runway = ws[RUNWAY_CELL].value or 0

    wb.close()

    return {
        "cash_balance": cash,
        "monthly_burn": burn,
        "runway_months": runway
    }


def get_budget_status():
    """Get budget vs actual for all categories."""
    wb = load_tracker()
    ws = wb.active

    status = []
    total_budget = 0
    total_actual = 0

    for category, row in BUDGET_CATEGORIES.items():
        budget = ws[f'B{row}'].value or 0
        actual = ws[f'C{row}'].value or 0
        variance = ws[f'D{row}'].value or 0

        total_budget += budget
        total_actual += actual

        pct = (actual / budget * 100) if budget > 0 else 0

        status.append({
            "category": category,
            "budget": budget,
            "actual": actual,
            "variance": variance,
            "percent": pct
        })

    wb.close()

    return {
        "categories": status,
        "total_budget": total_budget,
        "total_actual": total_actual
    }


def update_budget(category, new_amount):
    """Update budget for a category."""
    wb = load_tracker()
    ws = wb.active

    if category not in BUDGET_CATEGORIES:
        wb.close()
        raise ValueError(f"Unknown category: {category}. Valid: {', '.join(BUDGET_CATEGORIES.keys())}")

    row = BUDGET_CATEGORIES[category]
    old_budget = ws[f'B{row}'].value or 0
    ws[f'B{row}'] = new_amount

    wb.save(TRACKER_PATH)
    wb.close()

    return {
        "category": category,
        "old_budget": old_budget,
        "new_budget": new_amount
    }


def update_cash_balance(new_balance):
    """Directly update cash balance."""
    wb = load_tracker()
    ws = wb.active

    old_balance = ws[CASH_BALANCE_CELL].value or 0
    ws[CASH_BALANCE_CELL] = new_balance

    wb.save(TRACKER_PATH)
    wb.close()

    return {
        "old_balance": old_balance,
        "new_balance": new_balance
    }


if __name__ == "__main__":
    # Test basic operations
    print("Testing tracker operations...")

    try:
        status = get_status()
        cash = status['cash_balance'] if isinstance(status['cash_balance'], (int, float)) else 0
        burn = status['monthly_burn'] if isinstance(status['monthly_burn'], (int, float)) else 0
        runway = status['runway_months'] if isinstance(status['runway_months'], (int, float)) else 0

        print(f"✓ Cash: ${cash:,.0f}")
        print(f"✓ Burn: ${burn:,.0f}/mo")
        print(f"✓ Runway: {runway:.1f} months")

        budget = get_budget_status()
        print(f"\n✓ Budget loaded: {len(budget['categories'])} categories")
        print(f"  Total budget: ${budget['total_budget']:,.0f}")
        print(f"  Total actual: ${budget['total_actual']:,.0f}")

    except Exception as e:
        print(f"✗ Error: {e}")
